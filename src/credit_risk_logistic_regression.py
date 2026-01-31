"""
credit_risk_logistic_regression.py
----------------------------------
Credit Risk Predictive Analytics (Logistic Regression)

GitHub/portable version:
- Uses repo-relative paths (no hardcoded machine filepath)
- Reads:  data/CreditRiskData.xlsx  (sheets: "Past Loans", "Loan Applicants")
- Writes: outputs/
    - logit_coefficients_pvalues.csv
    - scored_loan_applicants.csv
    - scored_loan_applicants.xlsx  (RapidMiner-style table formatting)

Modeling notes:
- Logistic Regression via statsmodels.Logit
- Handles numeric + categorical predictors
    - numeric: median imputation
    - categorical: mode imputation + one-hot encoding
- Standardizes features for numerical stability, then back-transforms coefficients
  so "Coefficient_unscaled" is comparable to non-standardized units.

Run:
    pip install -r requirements.txt
    python src/credit_risk_logistic_regression.py
"""

from __future__ import annotations

from pathlib import Path
import warnings

import numpy as np
import pandas as pd
import statsmodels.api as sm

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule


# -----------------------------
# Repo paths (portable)
# -----------------------------
REPO_ROOT = Path(__file__).resolve().parents[1]          # repo_root/src/script.py -> repo_root
DATA_DIR = REPO_ROOT / "data"
OUT_DIR = REPO_ROOT / "outputs"
OUT_DIR.mkdir(exist_ok=True)

EXCEL_FILE = "CreditRiskData.xlsx"
FILE_PATH = DATA_DIR / EXCEL_FILE

SHEET_TRAIN = "Past Loans"
SHEET_SCORE = "Loan Applicants"

ID_COL = "Applicant ID"
LABEL_COL = "Good Risk"

# Flag as "uncertain" if the model's max class probability is below this threshold
UNCERTAIN_THRESHOLD = 0.70

# If you want the RapidMiner-style column names
PRED_COL = "prediction(Good Risk)"
CONF_YES_COL = "confidence(Yes)"
CONF_NO_COL = "confidence(No)"


# -----------------------------
# Helpers
# -----------------------------
def normalize_label(series: pd.Series) -> pd.Series:
    """Convert Yes/No-like labels to 1/0."""
    s = series.astype(str).str.strip().str.lower()
    mapped = s.map({
        "yes": 1, "no": 0,
        "y": 1, "n": 0,
        "true": 1, "false": 0,
        "1": 1, "0": 0
    })
    if mapped.isna().any():
        mapped = pd.to_numeric(series, errors="coerce")
    return mapped


def split_numeric_categorical(df: pd.DataFrame) -> tuple[list[str], list[str]]:
    """Return lists of numeric and categorical columns."""
    numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    categorical_cols = [c for c in df.columns if c not in numeric_cols]
    return numeric_cols, categorical_cols


def impute_and_encode(
    train_X: pd.DataFrame,
    score_X: pd.DataFrame
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Impute missing values and one-hot encode categorical columns
    in a way that keeps train/score columns aligned.
    """
    train = train_X.copy()
    score = score_X.copy()

    # Force consistent dtypes where possible
    # (Let pandas infer; we will handle numeric/categorical afterwards.)
    num_cols_train, cat_cols_train = split_numeric_categorical(train)

    # Numeric: median imputation
    if num_cols_train:
        med = train[num_cols_train].median(numeric_only=True)
        train[num_cols_train] = train[num_cols_train].fillna(med)
        score[num_cols_train] = score[num_cols_train].fillna(med)

    # Categorical: mode imputation (most frequent)
    if cat_cols_train:
        for c in cat_cols_train:
            mode_val = train[c].mode(dropna=True)
            fill_val = mode_val.iloc[0] if not mode_val.empty else ""
            train[c] = train[c].fillna(fill_val).astype(str)
            score[c] = score[c].fillna(fill_val).astype(str)

        # One-hot encode based on combined categories
        combined = pd.concat([train[cat_cols_train], score[cat_cols_train]], axis=0)
        combined_dum = pd.get_dummies(combined, columns=cat_cols_train, drop_first=True)

        train_dum = combined_dum.iloc[:len(train)].reset_index(drop=True)
        score_dum = combined_dum.iloc[len(train):].reset_index(drop=True)

        # Replace categorical columns with dummies
        train = pd.concat([train.drop(columns=cat_cols_train).reset_index(drop=True), train_dum], axis=1)
        score = pd.concat([score.drop(columns=cat_cols_train).reset_index(drop=True), score_dum], axis=1)

    # Ensure numeric
    train = train.apply(pd.to_numeric, errors="coerce")
    score = score.apply(pd.to_numeric, errors="coerce")

    # Any remaining NaN (from coercion) -> median
    med2 = train.median(numeric_only=True)
    train = train.fillna(med2)
    score = score.fillna(med2)

    return train, score


def standardize_train(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.Series, pd.Series]:
    """
    Standardize columns using training mean/std:
        x_scaled = (x - mean) / std
    Returns scaled df + mean + std.
    """
    mean = df.mean(axis=0)
    std = df.std(axis=0).replace(0, 1)  # avoid divide-by-zero
    scaled = (df - mean) / std
    return scaled, mean, std


def standardize_apply(df: pd.DataFrame, mean: pd.Series, std: pd.Series) -> pd.DataFrame:
    """Apply training mean/std to standardize scoring data."""
    # Reindex to be safe
    df = df.reindex(columns=mean.index, fill_value=0)
    return (df - mean) / std


def write_scored_excel(scored_df: pd.DataFrame, out_path: Path) -> None:
    """
    Create a RapidMiner-style Excel output:
    - Row No.
    - Applicant ID
    - prediction(Good Risk)
    - confidence(Yes)
    - confidence(No)
    - max_confidence
    - uncertain_flag
    With basic formatting + conditional color scale on max_confidence.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Scored Loan Applicants"

    # Build display table with "Row No." first
    display = scored_df.copy()
    display.insert(0, "Row No.", range(1, len(display) + 1))

    headers = display.columns.tolist()
    ws.append(headers)

    # Write rows
    for row in display.itertuples(index=False, name=None):
        ws.append(list(row))

    # Header styling
    header_font = Font(bold=True)
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Freeze header
    ws.freeze_panes = "A2"

    # Auto filter
    ws.auto_filter.ref = ws.dimensions

    # Column widths (reasonable defaults)
    for col_idx, h in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        # Wider for text columns, medium for numeric
        width = 14
        if h in ("Row No.",):
            width = 8
        elif h == ID_COL:
            width = 14
        elif "confidence" in h.lower():
            width = 16
        elif "prediction" in h.lower():
            width = 18
        elif h in ("uncertain_flag",):
            width = 14
        elif h in ("max_confidence",):
            width = 16
        ws.column_dimensions[col_letter].width = width

    # Number formats
    # Confidence columns as percentages with 2 decimals
    percent_cols = [c for c in headers if "confidence" in c.lower()]
    percent_cols += ["max_confidence"] if "max_confidence" in headers else []
    for col_name in percent_cols:
        col_idx = headers.index(col_name) + 1
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col_idx).number_format = "0.00%"

    # Conditional formatting on max_confidence (if present)
    if "max_confidence" in headers:
        col_idx = headers.index("max_confidence") + 1
        col_letter = get_column_letter(col_idx)
        data_range = f"{col_letter}2:{col_letter}{ws.max_row}"
        rule = ColorScaleRule(
            start_type="min", start_color="F8696B",   # red
            mid_type="percentile", mid_value=50, mid_color="FFEB84",  # yellow
            end_type="max", end_color="63BE7B"       # green
        )
        ws.conditional_formatting.add(data_range, rule)

    wb.save(out_path)


# -----------------------------
# Main workflow
# -----------------------------
def main() -> None:
    if not FILE_PATH.exists():
        raise FileNotFoundError(
            f"Missing dataset:\n  {FILE_PATH}\n\n"
            "Fix:\n"
            f"1) Create a folder named 'data' at your repo root\n"
            f"2) Put {EXCEL_FILE} inside /data\n"
        )

    # Load sheets
    past = pd.read_excel(FILE_PATH, sheet_name=SHEET_TRAIN)
    apps = pd.read_excel(FILE_PATH, sheet_name=SHEET_SCORE)

    print("Loaded:")
    print(" Past Loans:", past.shape)
    print(" Loan Applicants:", apps.shape)
    print()

    # Prepare y
    y = normalize_label(past[LABEL_COL])
    if y.isna().any():
        bad_rows = y[y.isna()].index[:10].tolist()
        raise ValueError(f"Label conversion failed (NaNs) at rows: {bad_rows}")

    # Separate predictors
    X_train_raw = past.drop(columns=[ID_COL, LABEL_COL])
    X_score_raw = apps.drop(columns=[ID_COL])

    # Impute + encode categoricals consistently
    X_train, X_score = impute_and_encode(X_train_raw, X_score_raw)

    # Standardize for stability
    X_train_scaled, x_mean, x_std = standardize_train(X_train)
    Xs_const = sm.add_constant(X_train_scaled, has_constant="add")

    # Fit model
    warnings.filterwarnings("ignore", category=RuntimeWarning)
    warnings.filterwarnings("ignore", category=UserWarning)

    logit = sm.Logit(y, Xs_const)

    try:
        result = logit.fit(disp=False, method="lbfgs", maxiter=5000)
    except Exception as e:
        # Common failure mode is perfect separation or quasi-separation
        raise RuntimeError(
            "Model fitting failed. This can happen with perfect/quasi separation "
            "or extreme multicollinearity.\n\n"
            f"Original error: {e}"
        ) from e

    converged = bool(getattr(result, "mle_retvals", {}).get("converged", True))
    if not converged:
        print("⚠️ Model did not fully converge. Interpret p-values cautiously.\n")
    else:
        print("✅ Model fit complete.\n")

    # Back-transform coefficients to original units (RapidMiner-comparable)
    params = result.params.copy()
    b0_scaled = params["const"]
    b_scaled = params.drop("const")

    # Unscaled coefficients for each feature
    b_unscaled = b_scaled / x_std

    # Unscaled intercept:
    # logit(p) = b0_scaled + Σ b_scaled * ((x - mean)/std)
    # => b0_unscaled = b0_scaled - Σ(b_scaled * mean/std)
    b0_unscaled = b0_scaled - np.sum(b_scaled * (x_mean / x_std))

    # Create coefficient table
    table = pd.DataFrame({
        "Attribute": ["Intercept"] + list(b_scaled.index),
        "Coefficient_unscaled": [b0_unscaled] + list(b_unscaled.values),
        "Coefficient_scaled": [b0_scaled] + list(b_scaled.values),
        "Std_Error_scaled": [result.bse["const"]] + list(result.bse.drop("const").values),
        "z_Value_scaled": [result.tvalues["const"]] + list(result.tvalues.drop("const").values),
        "p_Value": [result.pvalues["const"]] + list(result.pvalues.drop("const").values),
    })

    out_pvals = OUT_DIR / "logit_coefficients_pvalues.csv"
    table.to_csv(out_pvals, index=False)

    # Print strongest/weakest predictors (exclude intercept)
    predictors_only = table[table["Attribute"] != "Intercept"].copy()
    strongest = predictors_only.sort_values("p_Value").head(10)
    weakest = predictors_only.sort_values("p_Value", ascending=False).head(10)

    print("Top predictors by smallest p-values (strongest):")
    print(strongest[["Attribute", "Coefficient_unscaled", "p_Value"]].to_string(index=False))
    print()

    print("Weakest predictors by largest p-values (poorest):")
    print(weakest[["Attribute", "Coefficient_unscaled", "p_Value"]].to_string(index=False))
    print()

    # Score applicants
    X_score_scaled = standardize_apply(X_score, x_mean, x_std)
    Xas_const = sm.add_constant(X_score_scaled, has_constant="add")

    prob_yes = result.predict(Xas_const).values
    prob_no = 1 - prob_yes
    pred = (prob_yes >= 0.5).astype(int)

    scored = pd.DataFrame({
        ID_COL: apps[ID_COL],
        PRED_COL: np.where(pred == 1, "Yes", "No"),
        CONF_YES_COL: prob_yes,
        CONF_NO_COL: prob_no,
    })
    scored["max_confidence"] = scored[[CONF_YES_COL, CONF_NO_COL]].max(axis=1)
    scored["uncertain_flag"] = scored["max_confidence"] < UNCERTAIN_THRESHOLD

    # Save CSV
    out_scored_csv = OUT_DIR / "scored_loan_applicants.csv"
    scored.to_csv(out_scored_csv, index=False)

    # Save Excel (RapidMiner-style formatting)
    out_scored_xlsx = OUT_DIR / "scored_loan_applicants.xlsx"
    write_scored_excel(scored, out_scored_xlsx)

    # Summary print
    print("Prediction counts (Loan Applicants):")
    print(scored[PRED_COL].value_counts().to_string())
    print()

    print("Most uncertain predictions (top 10):")
    print(scored.sort_values("max_confidence").head(10).to_string(index=False))
    print()

    print("✅ Saved outputs:")
    print(f"- {out_scored_csv}")
    print(f"- {out_scored_xlsx}")
    print(f"- {out_pvals}")


if __name__ == "__main__":
    main()
